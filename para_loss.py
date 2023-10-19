# -*- coding: utf-8 -*-
"""
Author: Rakshit Joshi
Date Created: September 22, 2023
Organization: Gallagher Re, Munich

Description:
This Python script automates parametric loss calculations based on Modified Mercalli Intensity (MMI) data.

System Requirements:
- OSGeo4W (for Windows) or PyQGIS library (for Linux)
- Python packages: Selenium 4.11, Webdriver_manager 4.0, requests 2.31, openpyxl 3.1, csv 1.0, zipfile

These specific versions need not be used. However, certain combinations did not work.
A python script to check your environment and install necessary itemas is also provided (dependencies.py). 

Windows Setup Instructions:
1. Download and install OSGeo4W from https://www.qgis.org/en/site/forusers/download.html.
2. Open the OSGeo4W shell and run "python-qgis-ltr" to set up environmental variables.
3. Install the required Python packages. For SSL-related issues, follow the steps mentioned below.

A good resource for more information on OSGeo4W usage: "https://www.e-education.psu.edu/geog489/node/2294"

Note: The path tp qgis plugins sys.path.append('C:\\OSGeo4W\\apps\\qgis\\python\\plugins')
must be specified before importing the processing module. 

Linux Setup Instructions:
1. Install the PyQGIS library using 'python -m pip install qgis'.
2. Set the prefix path to the QGIS application before QGIS initialization (qgs.initQgis())
3. Download and install the required Python packages.

SSL Error Resolution:
If Module1 produces an SSL error when requesting data from USGS:
a) In the OSGeo4W shell, run the following commands:
   import requests as r
   print(r.certs.where())
b) Copy the certificates from the output address to "C:\\OSGeo4W\Apps\openssl\certs".

Running the Script:

0. Define the base folder, event ID, and the shapefile_path
1. Place this para_loss.py script in the OSGeo4W folder (for Windows users).
2. On Windows, open the OSGeo4W shell to set environment variables.
3. Run "python-qgis-ltr para_loss.py" to execute the script.

For any questions or issues, feel free to contact me.
"""




# Import all packages
import os
import zipfile
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
import time
import requests
import warnings
warnings.filterwarnings('ignore')
from qgis.core import *
import openpyxl
import csv
import sys
sys.path.append('C:\\OSGeo4W\\apps\\qgis\\python\\plugins')
import processing



# Initialzie your paths and eventid. Event id for an earthquake can be found on "https://earthquake.usgs.gov/" 

event_id = "us7000kufc"
base_folder = os.path.join("C:\\FSEC_Parametric_EQ_Loss_Calculation", f"Calculation_{event_id}")
shapefile_path = r'C:\\Morocco_Communes_Shapefile\\Morocco_Communes_Shapefile.shp'
original_excel_file = os.path.join(base_folder, "Calculation.xlsx")


# Module1: Download the ESRI raster files from  "https://earthquake.usgs.gov/"

def download_esri_raster_data(event_id, base_folder):
    # Define the download folder relative to the base folder
    download_folder = os.path.join(base_folder, f"ESRI_Raster_Files_{event_id}")

    # Make certain the download folder exists, create it if not
    if not os.path.exists(download_folder):
        os.makedirs(download_folder)

    # Configure Chrome driver and make it run in headless mode 
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument(f"--download-path={os.path.abspath(download_folder)}")
    driver = webdriver.Chrome(options=chrome_options)

    # Go to the webpage that contains ESRI raster files 
    driver.get(f'https://earthquake.usgs.gov/earthquakes/eventpage/{event_id}/shakemap/intensity')

    # Wait for elements to load
    while True:
        try:
            download_menu = driver.find_element('id', 'mat-expansion-panel-header-0')
        except NoSuchElementException:
            time.sleep(0.2)
            continue
        else:
            break

    download_menu.click()

    while True:
        try:
            downloads = driver.find_element('id', 'cdk-accordion-child-0')
        except NoSuchElementException:
            time.sleep(0.2)
            continue
        else:
            break

    
    links = downloads.find_elements('css selector', 'a')
    esri_raster_links = [link.get_attribute('href') for link in links if 'zip' in link.text.lower()]

    for file_url in esri_raster_links:
        file_name = file_url.split('/')[-1]  # Extract the file name from the URL

        # Send request to download the file
        response = requests.get(file_url)

        if response.status_code == 200:
            # Specify the full path to save the file in the download folder
            file_path = os.path.join(download_folder, file_name)
            with open(file_path, 'wb') as file:
                file.write(response.content)
            print(f"Downloaded {file_name}")

            # Unzip the downloaded file
            output_subfolder = os.path.join(download_folder, file_name.replace(".zip", ""))
            os.makedirs(output_subfolder, exist_ok=True)
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(output_subfolder)
            print(f"Unzipped {file_name}")

            # Remove the original ZIP file
            os.remove(file_path)
        else:
            print(f"Failed to download {file_name}")

    # Close the Driver
    driver.quit()


# Module2: Derive Maximum MMI per Commune from the USGS ShakeMap MMI data

def load_shakemapMMI_raster(event_id, base_folder):
    # Define the folder where the ShakeMap MMI data is downloaded
    mmi_data_folder = os.path.join(base_folder, "ESRI_Raster_Files_us7000kufc", "raster")

    # Import ShakeMap MMI data
    mmi_file = os.path.join(mmi_data_folder, 'mmi_mean.flt')

    # Add the MMI data as a raster layer
    mmi_layer_name = f'MMI_WGS84_{event_id}'
    mmi_output_folder = base_folder  # Replace with the event sub-folder path
    mmi_output_file = os.path.join(mmi_output_folder, f'{mmi_layer_name}.tif')

    mmi_layer = QgsRasterLayer(mmi_file, mmi_layer_name)
    if not mmi_layer.isValid():
        print(f"Failed to load MMI layer: {mmi_layer_name}")
    else:
        # Set the coordinates to EPSG:4326 - WGS 84
        crs = QgsCoordinateReferenceSystem.fromEpsgId(4326)
        mmi_layer.setCrs(crs)

        # Add the layer to the project
        project.addMapLayer(mmi_layer)
        print(f"Added MMI layer as: {mmi_layer}")

        pipe = QgsRasterPipe()
        pipe.set(mmi_layer.dataProvider())
        file_writer = QgsRasterFileWriter(mmi_output_file)
        file_writer.writeRaster(pipe, mmi_layer.width(), mmi_layer.height(), mmi_layer.extent(), mmi_layer.crs())
 
    project.write(project_path)
    return mmi_output_file

# Module3: Load and add commune shapefiles as vector layer
            
def  load_communes_vector(event_id, base_folder, shapefile_path):
        
        # Load the Communes shapefile
        #shapefile_path = os.path.join(base_folder, "Morocco_Communes_Shapefile", "Morocco_Communes_Shapefile.shp")
        
        communes_layer = QgsVectorLayer(shapefile_path, 'Communes', 'ogr')
        # Check if the layer was loaded successfully
        if not communes_layer.isValid():
            print('Error: Could not load Communes shapefile')
        else:
            # Set the CRS to EPSG:4326 - WGS 84
            crs = QgsCoordinateReferenceSystem.fromEpsgId(4326)
            communes_layer.setCrs(crs)
            project.addMapLayer(communes_layer)

            # Define the output path for the new shapefile
            output_folder =  base_folder
            output_filename = f'Communes_WGS84_{event_id}.shp'
            output_shapefile = os.path.join(output_folder, output_filename)

            #Write shapefile
            QgsVectorFileWriter.writeAsVectorFormat(communes_layer, output_shapefile, 'UTF-8', communes_layer.crs(), 'ESRI Shapefile')
            print(f'Shapefile saved as {output_shapefile}')

            # Add the saved file to the map
            saved_layer = QgsVectorLayer(output_shapefile, 'Communes_WGS84', 'ogr')
            project.addMapLayer(saved_layer)
           
            return output_shapefile

# Module 4: Run Zonal Statistics. The added maximimum value will be prefixed as column name "OUT_max". If you change this name >> also change name in function modify_calc() below.

def run_zonalstats(mmi_layer_path, communes_layer_path, event_id, base_folder):
    #Load all libraries. Note: For stand alone qgis applications, we need to use QgsNativeAlgorithms. Follow this import sequenc.
    print(event_id)
    from qgis.analysis import QgsNativeAlgorithms
    QgsApplication.processingRegistry().addProvider(QgsNativeAlgorithms())
    from processing.core.Processing import Processing

    Processing.initialize()

    # Define the output layer path for zonal statistics
    output_layer_path = base_folder  # Replace with the desired output path
    # Define parameters
    params = {'INPUT_RASTER': mmi_layer_path, 'RASTER_BAND': 1, 'INPUT_VECTOR': communes_layer_path,
              'COLUMN_PREFIX': 'OUT_', 'STATISTICS': 6,  # Use 2 for "max" statistic
              'OUTPUT': output_layer_path}

    processing.run("qgis:zonalstatistics", params)

    resulting_layer = QgsVectorLayer(output_layer_path, f"New_Communes_MaxMMI_{event_id}", "ogr")

    if resulting_layer.isValid():
        project.addMapLayer(resulting_layer)
        print("Zonal statistics layer added to the project.")
    else:
        print("Error: Failed to load the zonal statistics layer.")

    return resulting_layer


#Module 5:  Save zonal statistics results to a CSV file
def save_to_csv(resulting_layer, event_id, base_folder):
    if not resulting_layer.isValid():
        print("Error: Failed to load Commune layer.")
    else:
        field_names = [field.name() for field in resulting_layer.fields()]
        print(field_names)
        if "OUT_max" not in field_names:
            print("Error: The 'Max_MMI' field does not exist in the attribute table.")
        else:
            # Export the Commune layer as a CSV file
            output_csv_path = base_folder
            csv_filename = f'Communes_maxMMI_{event_id}.csv'
            output_csv_file = os.path.join(output_csv_path, csv_filename)
            # Define CSV export parameters
            csv_params = {
                'INPUT': resulting_layer,
                'OUTPUT': output_csv_file,
                'SEPARATOR': ',',
                'CREATE_CSVT': 'NO',
                'GEOMETRY': 'AS_WKT',
                'LINEFORMAT': 'Default',
                'WRITE_BOM': 'NO',
                'IF_AMBIGUOUS': 'YES',
                'FORCE_UTF8': 'YES',
                'DIALECT': 'UTF-8'
            }

            # Export the Commune layer to CSV
            processing.run("qgis:exportaddgeometrycolumns",
                           {'INPUT': resulting_layer, 'CALC_METHOD': 1, 'OUTPUT': output_csv_file})
            # processing.run("qgis:splitvectorlayer", {'INPUT': output_csv_file, 'FIELD': 'OUT_max', 'OUTPUT': output_csv_file})

            # Remove the additional geometry column
            with open(output_csv_file, 'r') as file:
                lines = file.readlines()
            with open(output_csv_file, 'w') as file:
                for line in lines:
                    if not line.startswith('geom'):
                        file.write(line)

            print(f"Commune data exported to CSV: {output_csv_file}")

# Module 6: Modify the "Calculations.xlxs" file and store results in a new file "Modified_Calculations.xlxs". This is the final results file. Check if the file is not broken.

def modify_calc():
    # Define the paths to the original and new Excel files
    original_excel_file = os.path.join(base_folder, "Calculation.xlsx")
    new_excel_file = os.path.join(base_folder, "Modified_Calculations.xlsx")  # Specify the new filename

    # Open the original Excel workbook
    workbook = openpyxl.load_workbook(original_excel_file)

    # Select the worksheet (tab) named "Data Input"
    worksheet = workbook['Data Input']

    # Open the CSV file for reading
    csv_file_path = os.path.join(base_folder, f'Communes_maxMMI_{event_id}.csv')
    with open(csv_file_path, 'r', newline='') as csv_file:
        # Use csv.DictReader to skip headers and access rows as dictionaries
        csv_reader = csv.DictReader(csv_file)

        # Start filling data from the third row
        row_index = 3

        # Iterate through rows in the CSV file and fill only the first two columns in the Excel worksheet
        for csv_row in csv_reader:
            worksheet.cell(row=row_index, column=1, value=csv_row['code_commu'])  # Replace 'code_commu' with the actual column name
            worksheet.cell(row=row_index, column=2, value=csv_row['OUT_max'])  # Replace 'OUT_max' with the actual column name
            row_index += 1  # Move to the next row in Excel

    # Save the modified Excel workbook as a new file
    workbook.save(new_excel_file)

    # Close the original workbook
    workbook.close()

    print(f"Modified Excel workbook saved as: {new_excel_file}")


# Initialize QGIS application, no need to set prefix path for QGS if run inside OSGeo4W (recommeneded).

qgs = QgsApplication([], True)
qgs.initQgis()
# Create new QGIS project
project = QgsProject.instance()
project_path = os.path.join(base_folder, f"project_{event_id}.qgs")
project.write(project_path)

# Run the functions and exit qgs >> if exit() command is crashing, check DLLs

download_esri_raster_data(event_id, base_folder)
a = load_shakemapMMI_raster(event_id, base_folder)
b = load_communes_vector(event_id, base_folder,shapefile_path)
c = run_zonalstats(a, b, event_id, base_folder)
save_to_csv(c,event_id, base_folder)
modify_calc()
#project.write()
qgs.exit()